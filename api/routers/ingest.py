"""Search-driven ingest router: search term → background scrape → embed."""

import asyncio
import csv
import hashlib
import io
import json
import os
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, UploadFile, File
from pydantic import BaseModel

from api.deps import get_conn, get_db_pool
from portfolio_architect.db.projects import get_project
from portfolio_architect.db.documents import insert_document, insert_chunks
from portfolio_architect.ingestion.chunker import chunk_text
from portfolio_architect.ingestion.fetchers import (
    pubmed_fetch, scholar_fetch, arxiv_fetch,
    record_to_text, deduplicate,
)

router = APIRouter(prefix="/projects/{project_id}", tags=["ingest"])

# In-memory progress stores keyed by project_id string
_progress: dict[str, dict] = {}          # ingest (search / files / DOI)
_analyze_progress: dict[str, dict] = {}  # analysis pipeline
_cancelled_projects: set[str] = set()    # projects whose build was cancelled


def request_cancel(project_id) -> None:
    """Signal the ingest/analysis pipeline for this project to stop at the next
    checkpoint. Also called before deleting a project so the background task
    stops writing to rows that are about to be removed."""
    pid = str(project_id)
    _cancelled_projects.add(pid)
    for store in (_progress, _analyze_progress):
        if pid in store:
            store[pid]["cancelled"] = True
            store[pid]["status"] = "cancelled"


def _is_cancelled(project_id) -> bool:
    return str(project_id) in _cancelled_projects


class SearchBody(BaseModel):
    query: str
    sources: list[str] = ["pubmed"]
    max_records: int = 200


@router.post("/search")
async def trigger_search(
    project_id: UUID,
    body: SearchBody,
    background_tasks: BackgroundTasks,
    conn=Depends(get_conn),
    pool=Depends(get_db_pool),
):
    project = await get_project(conn, project_id)
    if not project:
        raise HTTPException(404, f"Project {project_id} not found")

    pid_str = str(project_id)
    if _progress.get(pid_str, {}).get("status") == "running":
        raise HTTPException(409, "An ingest job is already running for this project.")

    _progress[pid_str] = {
        "status": "running",
        "query": body.query,
        "sources": body.sources,
        "fetched": 0,
        "ingested": 0,
        "embedded": 0,
        "total": 0,
        "error": None,
    }

    background_tasks.add_task(_run_ingest, pool, project_id, body)
    return {"status": "started", "project_id": pid_str}


@router.get("/ingest-status")
async def ingest_status(project_id: UUID, conn=Depends(get_conn)):
    pid_str = str(project_id)
    progress = _progress.get(pid_str, {"status": "idle"})

    # Also pull live DB counts
    doc_rows = await conn.fetch(
        "SELECT COUNT(*) AS cnt FROM documents WHERE project_id = ?", pid_str
    )
    emb_rows = await conn.fetch(
        "SELECT COUNT(*) AS cnt FROM documents WHERE project_id = ? AND embedded = 1", pid_str
    )
    pending_rows = await conn.fetch(
        "SELECT COUNT(*) AS cnt FROM documents WHERE project_id = ? AND embedded = 0", pid_str
    )

    return {
        **progress,
        "db_total": doc_rows[0]["cnt"] if doc_rows else 0,
        "db_embedded": emb_rows[0]["cnt"] if emb_rows else 0,
        "db_pending": pending_rows[0]["cnt"] if pending_rows else 0,
    }


async def _run_ingest(pool, project_id: UUID, body: SearchBody) -> None:
    pid_str = str(project_id)
    _cancelled_projects.discard(pid_str)
    prog = _progress[pid_str]

    def update(**kw):
        prog.update(kw)

    try:
        records: list[dict] = []
        share = body.max_records // max(len(body.sources), 1)

        if "pubmed" in body.sources:
            pubmed_recs = await asyncio.to_thread(pubmed_fetch, body.query, share)
            records.extend(pubmed_recs)
            update(fetched=len(records), total=body.max_records)

        if "scholar" in body.sources:
            scholar_recs = await asyncio.to_thread(scholar_fetch, body.query, share)
            records.extend(scholar_recs)
            update(fetched=len(records), total=body.max_records)

        if "arxiv" in body.sources:
            arxiv_recs = await asyncio.to_thread(arxiv_fetch, body.query, share)
            records.extend(arxiv_recs)
            update(fetched=len(records))

        records = deduplicate(records)
        if len(records) > body.max_records:
            records = records[: body.max_records]

        update(fetched=len(records), total=len(records))

        # Ingest into DB — insert_document returns existing row on duplicate (project_id, source_id)
        async with pool.acquire() as conn:
            new_count = 0
            for idx, rec in enumerate(records):
                raw_text = record_to_text(rec)
                chunks = chunk_text(raw_text)
                if not chunks:
                    continue
                doc = await insert_document(conn, project_id, rec["source_id"], raw_text)
                # Only chunk/embed if this is a fresh insert (chunk_count == 0)
                if not doc.get("chunk_count"):
                    await insert_chunks(conn, doc["id"], project_id, chunks)
                    new_count += 1
                if idx % 20 == 0:
                    update(ingested=idx + 1)
            update(ingested=new_count)

        update(ingested=len(records), status="embedding")

        # Trigger embedding for all unembedded chunks
        await _embed_all(pool, project_id)
        update(status="done", embedded=len(records))

    except Exception as e:
        update(status="error", error=str(e)[:300])


@router.post("/upload")
async def upload_dataset(
    project_id: UUID,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    pool=Depends(get_db_pool),
    conn=Depends(get_conn),
):
    """Accept a CSV or Excel file and ingest rows as documents."""
    project = await get_project(conn, project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    content = await file.read()
    records = await _records_from_upload(file.filename, content)
    if not records:
        raise HTTPException(
            400,
            "No usable records. Supported: .csv/.xlsx (with a title or abstract column) "
            "or .pdf (with extractable text).",
        )

    records = deduplicate(records)
    pid_str = str(project_id)
    _progress[pid_str] = {"status": "running", "fetched": len(records), "ingested": 0, "embedded": 0, "total": len(records), "error": None}
    background_tasks.add_task(_ingest_records, pool, project_id, records)
    return {"status": "started", "record_count": len(records)}


@router.post("/ingest-files")
async def ingest_files(
    project_id: UUID,
    background_tasks: BackgroundTasks,
    files: list[UploadFile] = File(...),
    pool=Depends(get_db_pool),
    conn=Depends(get_conn),
):
    """Ingest multiple files at once — a multi-file or whole-folder selection.
    Accepts .csv/.xlsx (rows need a title or abstract column) and .pdf (ingested
    by extracted text). Unreadable files are reported in `skipped`, not fatal."""
    if not await get_project(conn, project_id):
        raise HTTPException(404, "Project not found")

    records: list[dict] = []
    skipped: list[str] = []
    for f in files:
        content = await f.read()
        recs = await _records_from_upload(f.filename, content)
        if recs:
            records.extend(recs)
        else:
            skipped.append(f.filename)

    records = deduplicate(records)
    if not records:
        raise HTTPException(400, "No usable records found in the uploaded files.")

    pid_str = str(project_id)
    _progress[pid_str] = {
        "status": "running", "fetched": len(records), "ingested": 0,
        "embedded": 0, "total": len(records), "error": None, "skipped": skipped,
    }
    background_tasks.add_task(_ingest_records, pool, project_id, records)
    return {"status": "started", "record_count": len(records), "files": len(files), "skipped": skipped}


class DoiListBody(BaseModel):
    dois: list[str]


@router.post("/ingest-dois")
async def ingest_dois(
    project_id: UUID,
    body: DoiListBody,
    background_tasks: BackgroundTasks,
    pool=Depends(get_db_pool),
    conn=Depends(get_conn),
):
    """Resolve a list of DOIs (PubMed first, Crossref fallback) and ingest each
    as a document (chunk + embed). Unresolvable DOIs are skipped."""
    if not await get_project(conn, project_id):
        raise HTTPException(404, "Project not found")

    dois = [d.strip() for d in body.dois if d and d.strip()]
    if not dois:
        raise HTTPException(400, "Provide at least one DOI.")

    pid_str = str(project_id)
    _progress[pid_str] = {
        "status": "running", "fetched": 0, "ingested": 0,
        "embedded": 0, "total": len(dois), "error": None,
    }
    background_tasks.add_task(_run_doi_ingest, pool, project_id, dois)
    return {"status": "started", "doi_count": len(dois)}


async def _run_doi_ingest(pool, project_id: UUID, dois: list[str]) -> None:
    """Resolve each DOI to a record, then ingest as documents. Reuses the
    PubMed/Crossref resolvers from the add-by-DOI pipeline."""
    from portfolio_architect.claims.doi_ingest import (
        _resolve_via_pubmed, _resolve_via_crossref, _clean_doi,
    )
    prog = _progress[str(project_id)]
    records: list[dict] = []
    # Per-DOI outcome so the UI can show exactly which DOIs resolved and which
    # were skipped (and why) — no more silently dropping a third of the inputs.
    results: list[dict] = []
    for i, raw in enumerate(dois):
        d = _clean_doi(raw)
        try:
            rec = await asyncio.to_thread(lambda: _resolve_via_pubmed(d) or _resolve_via_crossref(d))
        except Exception:
            rec = None
        if rec and (rec.get("abstract") or "").strip():
            rec.setdefault("authors", "")
            rec.setdefault("url", "")
            records.append(rec)
            results.append({"doi": raw, "status": "resolved",
                            "title": rec.get("title"), "source_id": rec.get("source_id")})
        elif rec:
            results.append({"doi": raw, "status": "no_abstract", "title": rec.get("title")})
        else:
            results.append({"doi": raw, "status": "unresolved"})
        prog["fetched"] = i + 1
        prog["results"] = results

    records = deduplicate(records)
    # Mark resolved-but-deduplicated inputs (same paper submitted twice / already
    # matched another DOI) so the counts add up for the researcher.
    surviving = {r.get("source_id") for r in records}
    for res in results:
        if res["status"] == "resolved" and res.get("source_id") not in surviving:
            res["status"] = "duplicate"
    prog["results"] = results
    prog["resolved_count"] = len(records)
    prog["total"] = len(records)
    if not records:
        prog["status"] = "error"
        prog["error"] = "None of the DOIs could be resolved to an abstract (PubMed/Crossref)."
        return
    await _ingest_records(pool, project_id, records)


def _row_to_record(row: dict) -> dict | None:
    title = str(row.get("title") or row.get("Title") or "").strip()
    abstract = str(row.get("abstract") or row.get("Abstract") or row.get("summary") or "").strip()
    if not title and not abstract:
        return None
    sid = "upload:" + hashlib.md5((title + abstract).encode()).hexdigest()[:12]
    return {
        "source": "upload",
        "source_id": sid,
        "title": title,
        "abstract": abstract,
        "authors": str(row.get("authors") or row.get("author") or ""),
        "journal": str(row.get("journal") or row.get("venue") or ""),
        "year": str(row.get("year") or row.get("pub_year") or ""),
        "doi": str(row.get("doi") or row.get("DOI") or ""),
        "url": str(row.get("url") or row.get("link") or ""),
    }


def _records_from_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [r for r in (_row_to_record(row) for row in reader) if r]


def _records_from_excel(content: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(503, "openpyxl not installed — cannot read Excel files")
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = _row_to_record(dict(zip(headers, row)))
        if rec:
            out.append(rec)
    return out


def _pdf_to_record(filename: str, content: bytes) -> dict | None:
    """Extract a PDF's text into a single document record. Returns None for
    PDFs with no extractable text (e.g. scanned images — no OCR here)."""
    from pdfminer.high_level import extract_text
    try:
        text = (extract_text(io.BytesIO(content)) or "").strip()
    except Exception:
        return None
    if len(text) < 40:  # nothing meaningful extracted
        return None
    base = os.path.splitext(os.path.basename(filename or "document.pdf"))[0]
    title = (base.replace("_", " ").replace("-", " ").strip() or "Untitled PDF")[:200]
    sid = "pdf:" + hashlib.md5(((filename or "") + text[:400]).encode("utf-8", "ignore")).hexdigest()[:12]
    return {"source": "pdf", "source_id": sid, "title": title, "abstract": text,
            "authors": "", "journal": "", "year": "", "doi": "", "url": ""}


async def _records_from_upload(filename: str, content: bytes) -> list[dict]:
    """Parse one uploaded file into document records by extension.
    Supports .csv, .xlsx/.xls, .pdf. Unknown types → []."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _records_from_csv(content)
    if name.endswith((".xlsx", ".xls")):
        return _records_from_excel(content)
    if name.endswith(".pdf"):
        rec = await asyncio.to_thread(_pdf_to_record, filename, content)
        return [rec] if rec else []
    return []


async def _ingest_records(pool, project_id: UUID, records: list[dict]) -> None:
    pid_str = str(project_id)
    prog = _progress.get(pid_str, {})
    async with pool.acquire() as conn:
        new_count = 0
        for idx, rec in enumerate(records):
            raw_text = record_to_text(rec)
            chunks = chunk_text(raw_text)
            if not chunks:
                continue
            doc = await insert_document(conn, project_id, rec["source_id"], raw_text)
            if not doc.get("chunk_count"):
                await insert_chunks(conn, doc["id"], project_id, chunks)
                new_count += 1
            if idx % 20 == 0:
                prog["ingested"] = idx + 1
    prog["status"] = "embedding"
    await _embed_all(pool, project_id)
    prog["status"] = "done"
    prog["embedded"] = new_count


async def _embed_all(pool, project_id: UUID) -> None:
    """Embed all unembedded chunks for the project."""
    from portfolio_architect.embedding.client import embed_batch
    from portfolio_architect.db.documents import update_chunk_embedding

    pid_str = str(project_id)
    prog = _progress.get(pid_str, {})
    BATCH = 64

    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id, content FROM chunks WHERE project_id = ? AND embedding IS NULL",
            pid_str,
        )

    if not rows:
        return

    embedded_count = 0
    doc_ids: set[str] = set()

    for i in range(0, len(rows), BATCH):
        batch = rows[i: i + BATCH]
        texts = [r["content"] for r in batch]
        try:
            embeddings = await embed_batch(texts)
        except Exception:
            continue
        async with pool.acquire() as conn:
            for row, emb in zip(batch, embeddings):
                await update_chunk_embedding(conn, row["id"], emb)

            # mark docs embedded using a sub-query approach
            chunk_ids = [r["id"] for r in batch]
            for chunk_id in chunk_ids:
                doc_row = await conn.fetchrow(
                    "SELECT document_id FROM chunks WHERE id = ?", chunk_id
                )
                if doc_row:
                    doc_ids.add(doc_row["document_id"])

        embedded_count += len(batch)
        prog["embedded"] = embedded_count

    async with pool.acquire() as conn:
        for doc_id in doc_ids:
            await conn.execute(
                "UPDATE documents SET embedded = 1 WHERE id = ?", doc_id
            )


# ── Analysis pipeline: documents → claims → clusters → cited answers ──────────

@router.get("/analyze-status")
async def analyze_status(project_id: UUID, conn=Depends(get_conn)):
    pid = str(project_id)
    prog = _analyze_progress.get(pid, {"status": "idle"})
    counts = await conn.fetchrow(
        """SELECT
             (SELECT COUNT(*) FROM claims WHERE project_id = ?) AS claims,
             (SELECT COUNT(*) FROM claim_clusters WHERE project_id = ?) AS clusters,
             (SELECT COUNT(*) FROM claim_clusters WHERE project_id = ? AND answer IS NOT NULL) AS answered,
             (SELECT COUNT(*) FROM documents WHERE project_id = ? AND claims_extracted = 0) AS unprocessed""",
        pid, pid, pid, pid,
    )
    return {**prog, "db_claims": counts["claims"], "db_clusters": counts["clusters"],
            "db_answered": counts["answered"], "db_unprocessed": counts["unprocessed"]}


@router.post("/analyze")
async def trigger_analyze(
    project_id: UUID,
    background_tasks: BackgroundTasks,
    conn=Depends(get_conn),
    pool=Depends(get_db_pool),
):
    """Run the analysis pipeline over the project's documents: extract claims →
    embed → cluster → synthesize cited answers → lay out the map. Resumable
    (only unprocessed documents are extracted). Makes LLM + embedding calls."""
    if not await get_project(conn, project_id):
        raise HTTPException(404, "Project not found")
    pid = str(project_id)
    if _analyze_progress.get(pid, {}).get("status") == "running":
        raise HTTPException(409, "Analysis is already running for this project.")
    _analyze_progress[pid] = {
        "status": "running", "phase": "extracting", "error": None,
        "docs_total": 0, "docs_done": 0, "claims_extracted": 0,
        "clusters_built": 0, "conversations_total": 0, "conversations_done": 0,
    }
    background_tasks.add_task(_run_analyze, pool, project_id)
    return {"status": "started"}


@router.post("/cancel")
async def cancel_build(project_id: UUID, conn=Depends(get_conn)):
    """Stop the import/analysis for this project at the next checkpoint. Whatever
    finished so far is kept; the project is not deleted."""
    if not await get_project(conn, project_id):
        raise HTTPException(404, "Project not found")
    request_cancel(project_id)
    return {"status": "cancelled"}


async def _embed_claims_for_analyze(pool, project_id: UUID, prog: dict) -> None:
    from portfolio_architect.embedding.client import embed_batch
    from portfolio_architect.embedding import codec
    async with pool.acquire() as conn:
        rows = await conn.fetch(
            "SELECT id, claim_text FROM claims WHERE project_id = ? AND claim_embedding IS NULL",
            str(project_id),
        )
    for i in range(0, len(rows), 256):
        batch = rows[i:i + 256]
        try:
            vecs = await embed_batch([r["claim_text"] for r in batch])
        except Exception:
            continue
        async with pool.acquire() as conn:
            for r, v in zip(batch, vecs):
                await conn.execute("UPDATE claims SET claim_embedding = ? WHERE id = ?", codec.encode(v), r["id"])


async def _run_analyze(pool, project_id: UUID) -> None:
    """Full document→answer pipeline. Never raises — failures land on the
    progress record as status='error'."""
    from portfolio_architect.db.claims import (
        get_unprocessed_documents, insert_claims, mark_document_processed)
    from portfolio_architect.claims.extraction import run_one
    from portfolio_architect.claims.clustering import cluster_project_claims, add_singleton_clusters
    from portfolio_architect.db.claim_clusters import (
        reset_project_clusters, get_clusters_for_project, get_cluster_members, set_conversation)
    from portfolio_architect.claims.conversation import build_conversation

    _cancelled_projects.discard(str(project_id))
    prog = _analyze_progress[str(project_id)]

    def _stop() -> bool:
        if _is_cancelled(project_id):
            prog["status"] = "cancelled"
            return True
        return False

    try:
        # 1. Extract claims from unprocessed documents (bounded concurrency).
        prog["phase"] = "extracting"
        async with pool.acquire() as conn:
            docs = await get_unprocessed_documents(conn, project_id)
        prog["docs_total"] = len(docs)
        # Extraction is one hosted-LLM call per doc, pure I/O wait — the DB write
        # is trivial. 10 left throughput on the table; 25 roughly halves the
        # phase wall-clock for a new project without overrunning Token Factory.
        sem = asyncio.Semaphore(25)

        async def _extract(doc):
            if _is_cancelled(project_id):
                return
            async with sem, pool.acquire() as conn:
                try:
                    res = await run_one(conn, project_id, doc)
                    claims = res.get("claims") or []
                    if claims:
                        await insert_claims(conn, project_id, doc["id"], claims)
                        prog["claims_extracted"] += len(claims)
                    await mark_document_processed(conn, doc["id"], res.get("research_question"))
                except Exception:
                    pass
            prog["docs_done"] += 1

        await asyncio.gather(*[_extract(d) for d in docs])
        if _stop(): return

        # 2. Embed the new claims.
        prog["phase"] = "embedding"
        await _embed_claims_for_analyze(pool, project_id, prog)
        if _stop(): return

        # 3. Cluster (verified, PubMed-only), rebuilding from scratch.
        prog["phase"] = "clustering"
        async with pool.acquire() as conn:
            await reset_project_clusters(conn, project_id)
            multi = await cluster_project_claims(conn, project_id, verified_only=True, exclude_trials=True)
            singles = await add_singleton_clusters(conn, project_id, exclude_trials=True)
        prog["clusters_built"] = len(multi) + len(singles)
        if _stop(): return

        # 4. Answers are synthesized LAZILY now — on first cluster open, in
        #    workbench.cluster_detail — so the map appears as soon as clustering
        #    finishes instead of blocking on one LLM call per cluster. (The CLI
        #    build_conversations.py still pre-synthesizes the frozen demo corpus.)
        prog["conversations_total"] = 0
        if _stop(): return

        # 5. Lay out the map (PCA over cluster centroids → coord_x/coord_y).
        prog["phase"] = "finalizing"
        from api.routers import workbench as wb
        async with pool.acquire() as conn:
            clusters = await get_clusters_for_project(conn, project_id)
            cids = [c["id"] for c in clusters]
            centroids = await wb._load_cluster_centroids(conn, project_id, cids)
            coords = wb._project_2d(centroids)
            for cid, (x, y) in coords.items():
                await conn.execute("UPDATE claim_clusters SET coord_x = ?, coord_y = ? WHERE id = ?", x, y, cid)
        wb.invalidate_caches(str(project_id))

        prog["status"] = "done"
    except Exception as e:
        prog["status"] = "error"
        prog["error"] = str(e)[:300]
