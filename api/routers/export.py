"""Export router: validated screening decisions → CSV/Excel/Google Sheets."""

import csv
import io
import json
import os
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from api.deps import get_conn
from portfolio_architect.db.projects import get_project

router = APIRouter(prefix="/projects/{project_id}/export", tags=["export"])

def _safe_filename(name: str) -> str:
    """Strip non-ASCII and special chars so filenames are safe for HTTP Content-Disposition."""
    import re
    name = name.encode("ascii", errors="ignore").decode()
    return re.sub(r"[^\w\-. ]", "_", name).strip().replace(" ", "_")


_COLUMNS = [
    "source_id", "title", "authors", "journal", "year", "doi", "url",
    "human_label", "llm_label", "llm_confidence", "human_reason", "reason_code",
    "reviewer", "timestamp",
]


async def _get_decisions(conn, project_id: UUID) -> list[dict]:
    pid = str(project_id)
    rows = await conn.fetch(
        """
        SELECT
            d.source_id,
            d.raw_content,
            dec.human_label,
            dec.llm_label,
            dec.llm_confidence,
            dec.human_reason,
            dec.reason_code,
            dec.reviewer,
            dec.timestamp
        FROM decisions dec
        JOIN documents d ON d.id = dec.document_id
        WHERE dec.project_id = ?
        ORDER BY dec.timestamp
        """,
        pid,
    )

    results = []
    for r in rows:
        # Parse structured fields out of raw_content (Title: ...\nAuthors: ... format)
        meta = _parse_raw_content(r["raw_content"] or "")
        results.append({
            "source_id": r["source_id"],
            "title": meta.get("title", ""),
            "authors": meta.get("authors", ""),
            "journal": meta.get("journal", ""),
            "year": meta.get("year", ""),
            "doi": meta.get("doi", ""),
            "url": meta.get("url", ""),
            "human_label": r["human_label"],
            "llm_label": r["llm_label"] or "",
            "llm_confidence": round(r["llm_confidence"] * 100) if r["llm_confidence"] else "",
            "human_reason": r["human_reason"] or "",
            "reason_code": r["reason_code"] or "",
            "reviewer": r["reviewer"] or "",
            "timestamp": (r["timestamp"] or "")[:19],
        })
    return results


def _parse_raw_content(text: str) -> dict:
    meta = {}
    for line in text.splitlines():
        if ":" in line:
            key, _, val = line.partition(":")
            meta[key.strip().lower()] = val.strip()
    return meta


# ---------------------------------------------------------------------------
# CSV download
# ---------------------------------------------------------------------------

@router.get("/csv")
async def export_csv(project_id: UUID, conn=Depends(get_conn)):
    project = await get_project(conn, project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    rows = await _get_decisions(conn, project_id)
    if not rows:
        raise HTTPException(404, "No validated decisions to export")

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)

    filename = f"{_safe_filename(project['name'])}_validated.csv"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Excel download
# ---------------------------------------------------------------------------

@router.get("/excel")
async def export_excel(project_id: UUID, conn=Depends(get_conn)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(503, "openpyxl not installed")

    project = await get_project(conn, project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    rows = await _get_decisions(conn, project_id)
    if not rows:
        raise HTTPException(404, "No validated decisions to export")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validated Papers"

    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(bold=True, color="FFFFFF")

    headers = [c.replace("_", " ").title() for c in _COLUMNS]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, col in enumerate(_COLUMNS, 1):
            val = row.get(col, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            # Colour include/exclude rows
            if col == "human_label":
                if val == "include":
                    cell.fill = PatternFill("solid", fgColor="D1FAE5")
                elif val == "exclude":
                    cell.fill = PatternFill("solid", fgColor="FEE2E2")

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{_safe_filename(project['name'])}_validated.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Google Sheets push
# ---------------------------------------------------------------------------

class SheetsBody(BaseModel):
    spreadsheet_id: str          # existing sheet ID — or empty to create new
    sheet_name: str = "Validated Papers"
    credentials_json: str = ""   # service-account JSON as string; falls back to env GOOGLE_SHEETS_CREDS


@router.post("/google-sheets")
async def export_google_sheets(
    project_id: UUID,
    body: SheetsBody,
    conn=Depends(get_conn),
):
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        raise HTTPException(503, "gspread not installed")

    project = await get_project(conn, project_id)
    if not project:
        raise HTTPException(404, "Project not found")

    rows = await _get_decisions(conn, project_id)
    if not rows:
        raise HTTPException(404, "No validated decisions to export")

    # Load credentials
    creds_raw = body.credentials_json or os.getenv("GOOGLE_SHEETS_CREDS", "")
    if not creds_raw:
        raise HTTPException(400, "Google credentials not provided. Set GOOGLE_SHEETS_CREDS env var or pass credentials_json in the request body.")

    try:
        creds_dict = json.loads(creds_raw)
        creds = Credentials.from_service_account_info(
            creds_dict,
            scopes=["https://www.googleapis.com/auth/spreadsheets"],
        )
        gc = gspread.authorize(creds)
    except Exception as e:
        raise HTTPException(400, f"Invalid Google credentials: {e}")

    try:
        if body.spreadsheet_id:
            sh = gc.open_by_key(body.spreadsheet_id)
        else:
            sh = gc.create(f"{project['name']} — Validated Papers")
            sh.share(creds_dict.get("client_email", ""), perm_type="user", role="writer")
    except Exception as e:
        raise HTTPException(400, f"Could not open/create spreadsheet: {e}")

    # Get or create the target worksheet
    try:
        ws = sh.worksheet(body.sheet_name)
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=body.sheet_name, rows=len(rows) + 10, cols=len(_COLUMNS))

    header = [c.replace("_", " ").title() for c in _COLUMNS]
    data = [header] + [[str(row.get(c, "")) for c in _COLUMNS] for row in rows]
    ws.update(data, "A1")

    # Bold header row
    ws.format("A1:N1", {"textFormat": {"bold": True}, "backgroundColor": {"red": 0.12, "green": 0.23, "blue": 0.37}})

    sheet_url = f"https://docs.google.com/spreadsheets/d/{sh.id}"
    return {
        "status": "ok",
        "spreadsheet_id": sh.id,
        "sheet_url": sheet_url,
        "rows_written": len(rows),
    }
